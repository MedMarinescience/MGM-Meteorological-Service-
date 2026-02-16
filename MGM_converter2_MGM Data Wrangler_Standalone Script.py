"""
MGM Data Wrangler - Standalone Script
======================================
Parses Turkish Meteorological Service (MGM) Excel report files
and combines all parameters into one cleaned Excel workbook
with one sheet per station.

Usage:
    1. Set DATA_DIR below to your folder path
    2. Run:  python mgm_data_wrangler.py
    3. Output will be saved in the same folder as 'MGM_Station_Data_Cleaned.xlsx'

Requirements:
    pip install pandas openpyxl
"""

import pandas as pd
import numpy as np
import re
import os
import sys
import warnings
warnings.filterwarnings('ignore')

# Make console output robust on Windows terminals with legacy encodings.
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

# ╔══════════════════════════════════════════════════════════════════╗
# ║  CHANGE THIS PATH TO YOUR FOLDER                               ║
# ╚══════════════════════════════════════════════════════════════════╝
DATA_DIR = r"C:\user\data_examples"

# Output file will be saved in the same folder
OUTPUT_FILE = os.path.join(DATA_DIR, "MGM_Station_Data_Cleaned.xlsx")


# ── Parameter mapping: normalized filename phrase → clean column name ──────
# Add new entries here if you get more parameter files from MGM
PARAM_MAP = {
    "toplam yağış mm kg m² omgi":                        "Precip_OMGI_mm",
    "toplam yağış mm kg m² manuel":                      "Precip_Manuel_mm",
    "maksimum sıcaklık c":                               "Tmax_C",
    "minimum sıcaklık c":                                "Tmin_C",
    "ortalama sıcaklık c":                               "Tmean_C",
    "ortalama bulutluluk 8 okta":                        "Cloud_Cover_Okta",
    "fırtına kaydı olan günler":                         "Storm_Day",
    "kuvvetli rüzgar ve fırtına olan günler":            "Strong_Wind_Storm_Day",
    "maksimum rüzgar yönü ve hızı m sn":                 "Max_Wind_Dir_Speed",
    "kar su eş değeri manuel":                           "SWE_Manuel_mm",
    "mevcut kar yüksekliği cm manuel":                   "Snow_Depth_Manuel_cm",
    "ortalama kar yüksekliği cm":                        "Snow_Depth_Mean_cm",
    "ortalama yeni kar yüksekliği cm":                   "New_Snow_Mean_cm",
    "toplam açık yüzey buharlaşma mm":                   "Open_Evap_mm",
    "toplam buharlaşma evapotranspirasyon mm":           "ET_mm",
}


# ── Helper functions ─────────────────────────────────────────────

def sanitize_filename(filename):
    """
    Convert the actual filename (with spaces, parentheses, etc.)
    into the underscore format used in PARAM_MAP keys.

    Example:
        'Günlük Minimum Sıcaklık (C).xlsx'
        → 'Günlük_Minimum_Sıcaklık__C_'
    """
    name = os.path.splitext(filename)[0]          # remove .xlsx
    # Remove the leading date/ID prefix like '20250722794F-'
    name = re.sub(r'^[0-9A-Za-z]+-', '', name, count=1)
    # Normalize separators and symbols to spaces
    name = re.sub(r'[\(\)\[\]\{\}\.,;:!?\-\+_=/\\\*\|%°÷]', ' ', name)
    # Collapse repeated whitespace and lower case (Unicode-safe)
    name = re.sub(r'\s+', ' ', name).strip().casefold()
    return name


def identify_parameter(filename):
    """Match a filename to a parameter column name."""
    sanitized = sanitize_filename(filename)
    for key, col_name in PARAM_MAP.items():
        # Check if either normalized text contains the other
        # (handles extra terms, symbols, and spacing variations)
        if key in sanitized or sanitized in key:
            return col_name
    return None


def parse_wind_value(val):
    """Parse wind string like '53° 10.5' → (direction, speed)"""
    if pd.isna(val) or str(val).strip() == '':
        return np.nan, np.nan
    s = str(val).strip()
    m = re.match(r'(\d+(?:\.\d+)?)\s*°?\s+(\d+(?:\.\d+)?)', s)
    if m:
        return float(m.group(1)), float(m.group(2))
    return np.nan, np.nan


def parse_mgm_block(df_raw, start_row):
    """
    Parse one year×station block from the MGM report format.

    The block layout (starting from start_row):
        Row 0: Yıl: YYYY  İstasyon Adı/No: STATION_NAME/ID
        Row 1: (empty)
        Row 2: Parameter title
        Row 3: Gün/Ay  |  1  |  2  | ... | 12
        Row 4-34: Day 1-31, columns 2-13 = months 1-12
    """
    header_text = str(df_raw.iloc[start_row, 1])

    year_match = re.search(r'Yıl:\s*(\d{4})', header_text)
    station_match = re.search(r'İstasyon Adı/No:\s*(.+)', header_text)

    if not year_match or not station_match:
        return None, None, None

    year = int(year_match.group(1))
    station_raw = station_match.group(1).strip()

    # Data starts 4 rows after the header (day 1 = start_row + 4)
    data_start = start_row + 4
    records = []

    for day_offset in range(31):  # days 1-31
        row_idx = data_start + day_offset
        if row_idx >= len(df_raw):
            break

        day = day_offset + 1

        for month_idx in range(12):  # months 1-12
            month = month_idx + 1
            col = month_idx + 2  # columns 2..13

            val = df_raw.iloc[row_idx, col] if col < df_raw.shape[1] else np.nan

            # Skip invalid dates (e.g., Feb 30, Apr 31)
            try:
                date = pd.Timestamp(year=year, month=month, day=day)
            except ValueError:
                continue

            records.append((date, val))

    return station_raw, year, records


def parse_file(filepath):
    """
    Parse an entire MGM Excel file.
    Returns dict: station_name → DataFrame with Date + parameter column(s)
    """
    filename = os.path.basename(filepath)
    param_name = identify_parameter(filename)

    if param_name is None:
        print(f"  ⚠ Could not identify parameter for: {filename}")
        print(f"    Sanitized name: {sanitize_filename(filename)}")
        print(f"    → Skipping this file. You may need to add a new entry to PARAM_MAP.")
        return {}

    is_wind = (param_name == "Max_Wind_Dir_Speed")

    df_raw = pd.read_excel(filepath, sheet_name='Report', header=None)

    # Find all station header rows
    header_mask = df_raw[1].astype(str).str.contains('İstasyon', na=False)
    header_rows = df_raw.index[header_mask].tolist()

    station_data = {}

    for hr in header_rows:
        station_raw, year, records = parse_mgm_block(df_raw, hr)
        if station_raw is None or not records:
            continue

        if station_raw not in station_data:
            station_data[station_raw] = []

        for date, val in records:
            if is_wind:
                d, s = parse_wind_value(val)
                station_data[station_raw].append({
                    'Date': date,
                    'Max_Wind_Dir_deg': d,
                    'Max_Wind_Speed_mps': s
                })
            else:
                try:
                    v = float(val) if pd.notna(val) else np.nan
                except (ValueError, TypeError):
                    v = np.nan
                station_data[station_raw].append({
                    'Date': date,
                    param_name: v
                })

    # Convert to DataFrames
    result = {}
    for stn, recs in station_data.items():
        df = pd.DataFrame(recs)
        if not df.empty:
            df = df.groupby('Date', as_index=False).first()
            df = df.sort_values('Date').reset_index(drop=True)
        result[stn] = df

    return result


def make_sheet_name(station_raw):
    """Create an Excel-safe sheet name (max 31 characters)."""
    m = re.search(r'/(\d+)$', station_raw)
    station_id = m.group(1) if m else None

    name_part = re.sub(r'/\d+$', '', station_raw).strip()
    name_part = name_part.replace('/', '_').replace(' ', '_')

    # Remove characters not allowed in Excel sheet names
    for ch in ['[', ']', ':', '*', '?', '\\']:
        name_part = name_part.replace(ch, '')

    if station_id:
        max_name = 31 - len(station_id) - 1
        short_name = name_part[:max_name]
        return f"{short_name}_{station_id}"

    return name_part[:31]


# ── MAIN ─────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("  MGM Data Wrangler")
    print("=" * 70)
    print(f"\n  Data folder : {DATA_DIR}")
    print(f"  Output file : {OUTPUT_FILE}\n")

    # Find all MGM Excel files in the folder
    all_xlsx = sorted([
        f for f in os.listdir(DATA_DIR)
        if f.endswith('.xlsx')
        and not f.startswith('~$')
        and 'MGM_Station_Data_Cleaned' not in f
    ])

    # Filter to only files that match a known parameter
    mgm_files = []
    skipped = []
    for f in all_xlsx:
        param = identify_parameter(f)
        if param is not None:
            mgm_files.append(f)
        else:
            skipped.append(f)

    print(f"Found {len(mgm_files)} recognized MGM parameter files:\n")
    for f in mgm_files:
        param = identify_parameter(f)
        print(f"  ✓ {param:<25s} ← {f}")

    if skipped:
        print(f"\nSkipped {len(skipped)} unrecognized files:")
        for f in skipped:
            print(f"  ✗ {f}")

    if not mgm_files:
        print("\n  No MGM files found! Check your DATA_DIR path.")
        return

    # ── Parse all files ──
    print(f"\n{'─' * 70}")
    print("Parsing files...\n")

    all_station_data = {}  # station → [list of DataFrames]

    for filename in mgm_files:
        filepath = os.path.join(DATA_DIR, filename)
        print(f"  Parsing: {filename[:60]}...")
        file_data = parse_file(filepath)
        for stn, df in file_data.items():
            if stn not in all_station_data:
                all_station_data[stn] = []
            all_station_data[stn].append(df)
        print(f"           → {len(file_data)} stations")

    print(f"\n  Total unique stations: {len(all_station_data)}")

    # ── Merge all parameters per station ──
    print(f"\n{'─' * 70}")
    print("Merging parameters per station...\n")

    merged = {}
    for stn, dfs in all_station_data.items():
        result_df = dfs[0]
        for other_df in dfs[1:]:
            result_df = pd.merge(
                result_df, other_df, on='Date', how='outer',
                suffixes=('', '_dup')
            )
            # Handle duplicate columns from merge
            dup_cols = [c for c in result_df.columns if c.endswith('_dup')]
            for dc in dup_cols:
                orig = dc.replace('_dup', '')
                if orig in result_df.columns:
                    result_df[orig] = result_df[orig].combine_first(result_df[dc])
                result_df.drop(columns=[dc], inplace=True)

        result_df = result_df.sort_values('Date').reset_index(drop=True)
        merged[stn] = result_df

    # ── Ensure unique sheet names ──
    sheet_names = {}
    used_names = set()
    for stn in sorted(merged.keys()):
        base = make_sheet_name(stn)
        name = base
        counter = 2
        while name in used_names:
            suffix = f"_{counter}"
            name = base[:31 - len(suffix)] + suffix
            counter += 1
        used_names.add(name)
        sheet_names[stn] = name

    # ── Write Excel output ──
    print(f"Writing Excel workbook...\n")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    header_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Arial', size=10)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # ── Summary sheet ──
    ws_summary = wb.create_sheet("_SUMMARY")
    summary_headers = [
        "Station_Name", "Station_ID", "Sheet_Name",
        "Date_Start", "Date_End", "N_Records", "Parameters"
    ]
    for c, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    summary_row = 2
    for stn in sorted(merged.keys()):
        df = merged[stn]
        sname = sheet_names[stn]
        m = re.search(r'/(\d+)$', stn)
        station_id = m.group(1) if m else ""
        param_cols = [c for c in df.columns if c != 'Date']

        ws_summary.cell(row=summary_row, column=1, value=stn).font = data_font
        ws_summary.cell(row=summary_row, column=2, value=station_id).font = data_font
        ws_summary.cell(row=summary_row, column=3, value=sname).font = data_font
        ws_summary.cell(row=summary_row, column=4,
                        value=str(df['Date'].min().date()) if not df.empty else "").font = data_font
        ws_summary.cell(row=summary_row, column=5,
                        value=str(df['Date'].max().date()) if not df.empty else "").font = data_font
        ws_summary.cell(row=summary_row, column=6, value=len(df)).font = data_font
        ws_summary.cell(row=summary_row, column=7,
                        value=", ".join(param_cols)).font = data_font
        summary_row += 1

    for c in range(1, 8):
        ws_summary.column_dimensions[get_column_letter(c)].width = \
            [35, 12, 30, 12, 12, 10, 60][c - 1]

    # ── Station sheets ──
    station_count = 0
    for stn in sorted(merged.keys()):
        df = merged[stn]
        sname = sheet_names[stn]
        ws = wb.create_sheet(sname)

        # Station info row
        ws.cell(row=1, column=1, value=f"Station: {stn}").font = \
            Font(name='Arial', bold=True, size=11, color='2F5496')
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=min(len(df.columns), 8))

        # Headers
        cols = list(df.columns)
        for c_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=2, column=c_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data
        for r_idx, row in df.iterrows():
            for c_idx, col_name in enumerate(cols, 1):
                val = row[col_name]
                cell = ws.cell(row=r_idx + 3, column=c_idx)

                if col_name == 'Date' and pd.notna(val):
                    cell.value = val
                    cell.number_format = 'YYYY-MM-DD'
                elif pd.isna(val):
                    cell.value = None
                else:
                    cell.value = val

                cell.font = data_font
                cell.border = thin_border

        # Column widths
        for c_idx, col_name in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = max(len(col_name) + 2, 13)

        # Freeze top rows + date column
        ws.freeze_panes = 'B3'

        station_count += 1
        if station_count % 20 == 0:
            print(f"  Written {station_count} station sheets...")

    wb.save(OUTPUT_FILE)

    # ── Final summary ──
    print(f"\n{'=' * 70}")
    print(f"  DONE!")
    print(f"{'=' * 70}")
    print(f"  Output file  : {OUTPUT_FILE}")
    print(f"  Station sheets: {station_count}")
    print(f"  Parameters:")

    all_params = set()
    for df in merged.values():
        all_params.update([c for c in df.columns if c != 'Date'])

    for p in sorted(all_params):
        count = sum(1 for df in merged.values() if p in df.columns)
        print(f"    • {p:<25s} → {count} stations")

    all_dates = pd.concat([df['Date'] for df in merged.values() if not df.empty])
    print(f"\n  Date range: {all_dates.min().date()} to {all_dates.max().date()}")
    print(f"{'=' * 70}")


if __name__ == "__main__":

    main()
